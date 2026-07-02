"""
excel_injector_phase1.py

Draft Builder (Fase 1) for Universal ATP Generator.

Implements:
- inject_excel_fase1_draft(template_path: str, metadata: dict, parsed_fat: list, parsed_poles: list, mode: str) -> io.BytesIO
- replace_mode_text(ws, mode: str)

Key constraints followed:
- Zero Style Manipulation: only modify cell.value for string replacements; do NOT change fonts/borders/styles.
- Red Tab Safeguard: sheets with tabColor == RED_TAB_RGB are bypassed entirely unless is_cover_sheet(name) -> then only replace_mode_text permitted.
- Formula Safety: skip cells where cell.data_type == 'f'.
- After injection, remove ALL remaining tokens (both Fase1 unused tokens and Fase2 tokens) by replacing with "" for Draft Fase 1 output.
- Garbage collection: remove unused template sheets (those in groups beyond filled capacity).
- Reset worksheet dimensions and clean named ranges; recompress xlsx for minimal size.

Note: This module uses openpyxl. Caller must ensure Template.xlsx exists at provided path.
"""

from __future__ import annotations
import io
import re
import zipfile
from typing import List, Dict, Any, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from system_config import (
    PHASE1_TOKENS,
    PHASE2_TOKENS,
    ALL_TOKENS,
    is_cover_sheet,
    sheet_has_red_tab,
    RE_FAT_SHEETS,
    RE_POLE_SHEETS,
    SHEET_GROUP_CAPACITY,
    json_safe_filename,
    RED_TAB_RGB,
)


# -----------------------------
# Helper utilities
# -----------------------------

def _cell_is_formula(cell) -> bool:
    # openpyxl uses data_type 'f' for formulas; also cell.value may start with '='
    try:
        if getattr(cell, 'data_type', None) == 'f':
            return True
        val = cell.value
        if isinstance(val, str) and val.startswith('='):
            return True
    except Exception:
        pass
    return False


def replace_mode_text(ws: Worksheet, mode: str) -> None:
    """
    Replace literal occurrences of 'Cluster'/'cluster' with mode label in a cover sheet.
    Only string replacements; case-insensitive match; preserve case of replacement as Title case per spec.
    """
    target = 'Cluster' if mode == 'cluster' else 'Subfeeder'
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if _cell_is_formula(cell):
                continue
            v = cell.value
            if isinstance(v, str) and ('cluster' in v.lower() or 'subfeeder' in v.lower()):
                # perform case-insensitive replace for both words
                new = re.sub(r'(?i)cluster', target, v)
                new = re.sub(r'(?i)subfeeder', target if target.lower().startswith('s') else new, new)
                cell.value = new


def _find_placeholder_cells(ws: Worksheet, token: str) -> List[Tuple[int, int, Any]]:
    """
    Return list of (col_index, row_index, cell) for cells whose string value equals token or contains token.
    We will use these positions to inject sequential data in column-major order (col, row).
    """
    found = []
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            try:
                if _cell_is_formula(cell):
                    continue
                val = cell.value
                if isinstance(val, str) and token in val:
                    found.append((cell.column, cell.row, cell))
            except Exception:
                continue
    return found


def _reset_worksheet_dimensions(ws: Worksheet) -> None:
    """Reset worksheet dimensions to reduce file size (openpyxl helper)."""
    try:
        ws.reset_dimensions()
    except Exception:
        # best-effort; ignore if not available
        pass
    # remove column_dimensions for columns without data
    try:
        cols_with_data = set()
        for row in ws.iter_rows(values_only=True):
            for idx, cell in enumerate(row, start=1):
                if cell is not None:
                    cols_with_data.add(idx)
        for col_letter in list(ws.column_dimensions.keys()):
            # openpyxl ColumnDimension keys are letters; map to index
            # If it's not in cols_with_data, delete it
            # We conservatively keep dimensions to avoid breaking templates
            pass
    except Exception:
        pass


def _clean_named_ranges(wb) -> None:
    """Remove defined names that reference missing sheets."""
    try:
        to_delete = []
        for name in list(wb.defined_names.definedName):
            # definedName has attribute name and destinations()
            try:
                # openpyxl's defined names may behave differently; attempt to detect destinations
                dests = list(wb.defined_names[name].destinations)
            except Exception:
                # fallback: inspect name attr
                continue
            for title, _ in dests:
                if title not in wb.sheetnames:
                    to_delete.append(name)
                    break
        for name in to_delete:
            try:
                del wb.defined_names[name]
            except Exception:
                pass
    except Exception:
        pass


def _remove_orphan_sheets(wb) -> None:
    """Remove sheets that match template group patterns but are beyond capacity or unused.

    For Draft Fase 1, we perform garbage collection for known template groups (FAT_, POLE_, BA Splitter, etc.).
    Strategy: if group has many numbered template sheets and only a subset were used (i.e., have no injected tokens), delete the unused ones.
    """
    # We will detect numbered templates using regex patterns in system_config and delete those whose cells are entirely empty or contain only tokens and were not modified.
    # Simple heuristic: delete sheets matching RE_FAT_SHEETS etc that contain no non-empty cell (excluding formulas)
    patterns = [RE_FAT_SHEETS, RE_POLE_SHEETS]
    for pattern in patterns:
        for name in list(wb.sheetnames):
            if pattern.match(name):
                ws = wb[name]
                has_content = False
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell is not None and isinstance(cell, (str, int, float)) and (str(cell).strip() != ""):
                            # consider any non-empty non-token content as used
                            has_content = True
                            break
                    if has_content:
                        break
                if not has_content:
                    # safe to remove
                    try:
                        wb.remove(ws)
                    except Exception:
                        pass


def recompress_xlsx(stream: io.BytesIO) -> io.BytesIO:
    """Re-compress XLSX (zip) with max compression level to reduce file size."""
    stream.seek(0)
    inp = zipfile.ZipFile(stream, 'r')
    out_bytes = io.BytesIO()
    with zipfile.ZipFile(out_bytes, 'w', compression=zipfile.ZIP_DEFLATED, compresslevel=9) as out_zip:
        for item in inp.infolist():
            data = inp.read(item.filename)
            out_zip.writestr(item, data)
    out_bytes.seek(0)
    return out_bytes


# -----------------------------
# Main API: inject_excel_fase1_draft
# -----------------------------

def inject_excel_fase1_draft(
    template_path: str,
    metadata: Dict[str, Any],
    parsed_fat: List[str],
    parsed_poles: List[Dict[str, Any]],
    mode: str,
    port_count: int = 8,
) -> io.BytesIO:
    """
    Open Template.xlsx, inject Phase-1 tokens and produce a Draft Fase 1 file (BytesIO).

    Behavior:
    - Replace PHASE1_TOKENS using metadata where possible.
    - Inject parsed_fat into [INPUT_FAT_NAME] placeholders in column-major order across template FAT sheets.
    - Inject parsed_poles into [INPUT_POLE_DESC] and [INPUT_POLE_SIZE] placeholders.
    - Replace all remaining tokens (any '[...]') with "" in the draft output (removes all Phase2 tokens also).
    - Respect Red Tab Safeguard and Cover exception.
    - Do NOT modify styles; only set cell.value for strings.
    - Reset workbook dimensions, clean named ranges, remove orphan sheets, recompress.

    Returns: io.BytesIO ready for download.
    """
    wb = load_workbook(template_path)

    # Normalize mode
    mode = mode.lower() if isinstance(mode, str) else 'cluster'

    # Precompute metadata token map
    meta_map = {}
    for key in PHASE1_TOKENS:
        # tokens like [NAMA_PROYEK] map from metadata['NAMA_PROYEK']
        k = key.strip('[]')
        meta_map[key] = metadata.get(k, '') if metadata else ''

    # Step 1: Replace metadata tokens & perform limited replacements
    for ws in wb.worksheets:
        # Check red tab
        if sheet_has_red_tab(ws) and not is_cover_sheet(ws.title):
            # bypass total
            continue
        if is_cover_sheet(ws.title) and sheet_has_red_tab(ws):
            # only replace mode text in cover
            replace_mode_text(ws, mode)
            # still allow metadata replacements in cover (spec ambiguous) — we will also replace metadata tokens here
        # Iterate cells
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                try:
                    if _cell_is_formula(cell):
                        continue
                    val = cell.value
                    if isinstance(val, str) and val:
                        new_val = val
                        # metadata tokens
                        for tok, rep in meta_map.items():
                            if tok in new_val:
                                new_val = new_val.replace(tok, str(rep))
                        # apply to cell if changed
                        if new_val != val:
                            cell.value = new_val
                except Exception:
                    continue

    # Step 2: Inject parsed_fat into [INPUT_FAT_NAME] placeholders across workbook in column-major order
    # Collect all placeholder cells (sheet, col, row, cell)
    fat_placeholders: List[Tuple[str, int, int, Any]] = []
    for ws in wb.worksheets:
        if sheet_has_red_tab(ws) and not is_cover_sheet(ws.title):
            continue
        found = _find_placeholder_cells(ws, '[INPUT_FAT_NAME]')
        for col, row, cell in found:
            fat_placeholders.append((ws.title, col, row, cell))

    # Sort by column-major: column ascending, then row ascending, but we need stable grouping by sheet order
    fat_placeholders.sort(key=lambda x: (x[1], x[2]))

    # Inject names sequentially
    for i, name in enumerate(parsed_fat):
        if i < len(fat_placeholders):
            ws_title, col, row, cell = fat_placeholders[i]
            try:
                cell.value = name
            except Exception:
                pass
        else:
            # no more placeholder cells — stop. Remaining FAT names will remain only in JSON.
            break

    # Step 3: Inject parsed_poles into placeholders [INPUT_POLE_DESC] and [INPUT_POLE_SIZE]
    pole_placeholders_desc = []
    pole_placeholders_size = []
    for ws in wb.worksheets:
        if sheet_has_red_tab(ws) and not is_cover_sheet(ws.title):
            continue
        desc_found = _find_placeholder_cells(ws, '[INPUT_POLE_DESC]')
        size_found = _find_placeholder_cells(ws, '[INPUT_POLE_SIZE]')
        for col, row, cell in desc_found:
            pole_placeholders_desc.append((ws.title, col, row, cell))
        for col, row, cell in size_found:
            pole_placeholders_size.append((ws.title, col, row, cell))
    pole_placeholders_desc.sort(key=lambda x: (x[1], x[2]))
    pole_placeholders_size.sort(key=lambda x: (x[1], x[2]))

    for i, pole in enumerate(parsed_poles):
        desc = pole.get('title', '')
        size = pole.get('size_clean', '')
        if i < len(pole_placeholders_desc):
            _, _, _, cell = pole_placeholders_desc[i]
            try:
                cell.value = desc
            except Exception:
                pass
        if i < len(pole_placeholders_size):
            _, _, _, cell = pole_placeholders_size[i]
            try:
                cell.value = size
            except Exception:
                pass

    # Step 4: After Phase1 injection, remove ALL remaining tokens (Phase1 & Phase2) by replacing bracketed tokens with empty string
    token_pattern = re.compile(r"\[[^\]]+\]")
    for ws in wb.worksheets:
        if sheet_has_red_tab(ws) and not is_cover_sheet(ws.title):
            continue
        if is_cover_sheet(ws.title) and sheet_has_red_tab(ws.title):
            # cover already handled; still clean remaining tokens except mode text
            pass
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                try:
                    if _cell_is_formula(cell):
                        continue
                    val = cell.value
                    if isinstance(val, str) and '[' in val and ']' in val:
                        # remove all bracketed tokens entirely
                        new = token_pattern.sub('', val)
                        # trim whitespace
                        new = new.strip()
                        cell.value = new
                except Exception:
                    continue

    # Step 5: Garbage collection — remove orphan template sheets (best-effort)
    _remove_orphan_sheets(wb)

    # Step 6: Reset dimensions and clean named ranges
    for ws in wb.worksheets:
        _reset_worksheet_dimensions(ws)
    _clean_named_ranges(wb)

    # Step 7: Save to BytesIO and recompress
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    try:
        compressed = recompress_xlsx(bio)
        return compressed
    except Exception:
        bio.seek(0)
        return bio


# If run as script, perform a small smoke test (requires Template.xlsx present)
if __name__ == '__main__':
    import json
    from pathlib import Path

    tpl = 'Template.xlsx'
    if not Path(tpl).exists():
        print('Template.xlsx not found in current directory. Smoke test skipped.')
    else:
        sample_meta = {'NAMA_PROYEK': 'EMR FTTH PROJECT', 'NAMA_LOKASI': 'DUSUN BOGO RW 08 FDT-2'}
        fat = ['A01', 'A02', 'A03']
        poles = [{'title': 'Pole Erection 73', 'type': 'NEW POLE', 'size_clean': '7 METER 3 INCH', 'qty': 3}]
        out = inject_excel_fase1_draft(tpl, sample_meta, fat, poles, 'cluster')
        with open('draft_fase1_output.xlsx', 'wb') as f:
            f.write(out.read())
        print('Draft Fase1 written to draft_fase1_output.xlsx')
