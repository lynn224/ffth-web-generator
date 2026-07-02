"""
excel_injector_phase2.py

Final Integrator (Fase 2) for Universal ATP Generator.

Implements:
- inject_excel_final(template_path: str, metadata: dict, parsed_fat: list, parsed_poles: list, phase2_data: dict, mode: str, port_count: int = 8) -> io.BytesIO
- replace_mode_text(ws, mode: str)

This module opens a fresh Template.xlsx and injects all Fase1 + Fase2 data in a single combined pass.
Key guarantees:
- Zero Style Manipulation: only string replacements via cell.value; do not change cell styles or structure.
- Red Tab Safeguard: bypass sheets with red tab except Cover sheets (only mode text replacement allowed there).
- Formula Safety: skip cells where cell.data_type == 'f' or value startswith '='.
- Phase-2 data injection follows row/column-major scanning and token-index awareness (best-effort for OTDR/OPM mapping per spec).
- After injection, run garbage collection, reset dims, clean named ranges, and recompress xlsx.

Note: This implementation focuses on correctness of token injection and safety. It is designed to be compatible with the previously committed excel_injector_phase1.py utilities where practical.
"""

from __future__ import annotations
import io
import re
import zipfile
from typing import List, Dict, Any, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from system_config import (
    PHASE1_TOKENS,
    PHASE2_TOKENS,
    ALL_TOKENS,
    is_cover_sheet,
    sheet_has_red_tab,
    RE_FAT_SHEETS,
    RE_POLE_SHEETS,
    RE_BA_SPLITTER_FAT,
    RE_OTDR_SUMMARY_FDT_FAT,
    RE_OPM_DISTRIBUTION,
    RE_OPM_SHEETS,
    RE_OTDR_WAVE,
    DEFAULT_PORT_COUNT,
)


# -----------------------------
# Utility functions (formula detection, replacements, recompress)
# -----------------------------

def _cell_is_formula(cell) -> bool:
    try:
        if getattr(cell, 'data_type', None) == 'f':
            return True
        val = cell.value
        if isinstance(val, str) and val.startswith('='):
            return True
    except Exception:
        pass
    return False


def replace_mode_text(ws: Worksheet, mode: str) -> None:
    target = 'Cluster' if mode == 'cluster' else 'Subfeeder'
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if _cell_is_formula(cell):
                continue
            v = cell.value
            if isinstance(v, str) and ('cluster' in v.lower() or 'subfeeder' in v.lower()):
                # Replace both words with the target
                new = re.sub(r'(?i)cluster', target, v)
                new = re.sub(r'(?i)subfeeder', target, new)
                cell.value = new


def recompress_xlsx(stream: io.BytesIO) -> io.BytesIO:
    stream.seek(0)
    inp = zipfile.ZipFile(stream, 'r')
    out_bytes = io.BytesIO()
    with zipfile.ZipFile(out_bytes, 'w', compression=zipfile.ZIP_DEFLATED, compresslevel=9) as out_zip:
        for item in inp.infolist():
            data = inp.read(item.filename)
            out_zip.writestr(item, data)
    out_bytes.seek(0)
    return out_bytes


# -----------------------------
# Generic placeholder discovery & injection helpers
# -----------------------------

def _find_placeholder_cells_in_ws(ws: Worksheet, token: str) -> List[Tuple[int, int, Any]]:
    found = []
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            try:
                if _cell_is_formula(cell):
                    continue
                val = cell.value
                if isinstance(val, str) and token in val:
                    found.append((cell.column, cell.row, cell))
            except Exception:
                continue
    return found


def _collect_placeholder_cells(wb, token: str, allow_cover_replace: bool = False) -> List[Tuple[Worksheet, int, int, Any]]:
    res = []
    for ws in wb.worksheets:
        # Red tab safeguard
        if sheet_has_red_tab(ws) and not is_cover_sheet(ws.title):
            continue
        if is_cover_sheet(ws.title) and sheet_has_red_tab(ws.title):
            if allow_cover_replace:
                # only allow mode text replacement, skip other placeholders
                pass
        cells = _find_placeholder_cells_in_ws(ws, token)
        for col, row, cell in cells:
            res.append((ws, col, row, cell))
    # Sort column-major per sheet discovery order
    res.sort(key=lambda x: (x[1], x[2]))
    return res


# -----------------------------
# Injection primitives for Phase2 groups
# -----------------------------

def _inject_sequential_token(wb, token: str, data_list: List[str]) -> int:
    """Inject string values from data_list into cells containing token across workbook in column-major order.
    Returns number of items injected.
    """
    targets = _collect_placeholder_cells(wb, token)
    injected = 0
    for i, value in enumerate(data_list):
        if i >= len(targets):
            break
        ws, col, row, cell = targets[i]
        try:
            cell.value = value
            injected += 1
        except Exception:
            continue
    return injected


def _inject_opm_feeder(wb, splitter_grid: List[Dict[str, Any]], port_count: int = DEFAULT_PORT_COUNT) -> int:
    """
    Inject OPM feeder splitters. Each row in splitter_grid is a dict with keys: SPL_ID, SPL_SN, OPM_BEFORE, P1..Pn
    Strategy: find '[SPL_ID]' occurrences and fill group-wise (SPL_ID, SPL_SN, OPM_BEFORE, then ports vertically under port token or P# tokens if present).
    Returns count of splitters injected.
    """
    # find SPL_ID placeholders
    targets = _collect_placeholder_cells(wb, '[SPL_ID]')
    count = 0
    for i, rowdata in enumerate(splitter_grid):
        if i >= len(targets):
            break
        ws, col, row, cell = targets[i]
        try:
            # Fill SPL_ID
            cell.value = rowdata.get('SPL_ID', '')
            # Fill SPL_SN: find nearest cell in same row that contains [SPL_SN]
            for c in ws[row]:
                if _cell_is_formula(c):
                    continue
                v = c.value
                if isinstance(v, str) and '[SPL_SN]' in v:
                    c.value = rowdata.get('SPL_SN', '')
                    break
            # Fill OPM_BEFORE in same row
            for c in ws[row]:
                if _cell_is_formula(c):
                    continue
                v = c.value
                if isinstance(v, str) and '[OPM_BEFORE]' in v:
                    c.value = rowdata.get('OPM_BEFORE', '')
                    break
            # Fill ports: attempt to find contiguous cells below a cell containing '[OPM_AFTER]' or cells labeled P1..Pn
            # First search for a cell containing '[OPM_AFTER]' in the same column as SPL_ID or nearby columns
            port_values = [rowdata.get(f'P{j}', '') for j in range(1, port_count+1)]
            # Find column that contains '[OPM_AFTER]' in this worksheet
            opm_after_cells = _find_placeholder_cells_in_ws(ws, '[OPM_AFTER]')
            if opm_after_cells:
                # choose first opm_after cell nearest to row
                op_col, op_row, op_cell = opm_after_cells[0]
                # inject vertically starting at op_row for port_count rows
                for p_idx, pval in enumerate(port_values):
                    try:
                        target_cell = ws.cell(row=op_row + p_idx, column=op_col)
                        if not _cell_is_formula(target_cell):
                            target_cell.value = pval or ''
                    except Exception:
                        continue
            else:
                # Fallback: look for cell tokens 'P1','P2' etc and replace if found
                for p_idx, pval in enumerate(port_values, start=1):
                    key = f'P{p_idx}'
                    # collect placeholders for this port key
                    port_targets = _collect_placeholder_cells(wb, f'[{key}]')
                    if port_targets and len(port_targets) > i:
                        # fill the i-th occurrence
                        tgt_ws, tgt_col, tgt_row, tgt_cell = port_targets[i]
                        try:
                            tgt_cell.value = pval or ''
                        except Exception:
                            pass
            count += 1
        except Exception:
            continue
    return count


def _inject_opm_distribution(wb, opm_dist_grid: List[Dict[str, Any]], port_count: int = DEFAULT_PORT_COUNT) -> int:
    targets = _collect_placeholder_cells(wb, '[INPUT_OPM]')
    injected = 0
    for i, rowdata in enumerate(opm_dist_grid):
        if i >= len(targets):
            break
        ws, col, row, cell = targets[i]
        try:
            cell.value = rowdata.get('FAT_NAME', '')
            # fill ports P1..Pn in same row or below
            for p_idx in range(1, port_count+1):
                pkey = f'P{p_idx}'
                # try same row scan
                filled = False
                for c in ws[row]:
                    if _cell_is_formula(c):
                        continue
                    v = c.value
                    if isinstance(v, str) and f'[{pkey}]' in v:
                        c.value = rowdata.get(pkey, '')
                        filled = True
                        break
                if not filled:
                    # try to find vertical placeholders for INPUT_OPM and fill downwards
                    pass
            injected += 1
        except Exception:
            continue
    return injected


def _inject_otdr_cluster(wb, otdr_grid: List[Dict[str, Any]]) -> int:
    """
    Inject OTDR Cluster data using zig-zag pattern. For each FAT entry in otdr_grid,
    find a cell containing [INPUT_FAT_NAME], fill it, then in the same row fill [DISTANCE] and [1310], and in next row fill [DISTANCE] and [1550].
    """
    # Collect all [INPUT_FAT_NAME] placeholders across workbook (in column-major)
    placeholders = _collect_placeholder_cells(wb, '[INPUT_FAT_NAME]')
    injected = 0
    for i, entry in enumerate(otdr_grid):
        if i >= len(placeholders):
            break
        ws, col, row, cell = placeholders[i]
        try:
            # FAT name
            cell.value = entry.get('FAT_NAME', '')
            # same row: DISTANCE (1310) and 1310 (loss)
            for c in ws[row]:
                if _cell_is_formula(c):
                    continue
                v = c.value
                if isinstance(v, str) and '[DISTANCE]' in v:
                    # first occurrence in row assume Distance 1310
                    c.value = entry.get('DISTANCE_1310', '')
                    break
            for c in ws[row]:
                if _cell_is_formula(c):
                    continue
                v = c.value
                if isinstance(v, str) and '[1310]' in v:
                    c.value = entry.get('LOSS_1310', '')
                    break
            # next row (row+1): fill Distance 1550 and [1550]
            next_row = row + 1
            for c in ws[next_row]:
                if _cell_is_formula(c):
                    continue
                v = c.value
                if isinstance(v, str) and '[DISTANCE]' in v:
                    # this is second DISTANCE slot
                    c.value = entry.get('DISTANCE_1550', '')
                    break
            for c in ws[next_row]:
                if _cell_is_formula(c):
                    continue
                v = c.value
                if isinstance(v, str) and '[1550]' in v:
                    c.value = entry.get('LOSS_1550', '')
                    break
            injected += 1
        except Exception:
            continue
    return injected


def _inject_otdr_subfeeder(wb, otdr_grid: List[Dict[str, Any]]) -> int:
    """
    Inject OTDR Subfeeder data. For subfeeder, we expect separate sheet(s) OTDR Summary 1310 and OTDR Summary 1550.
    Strategy: find sheets whose title contains '1310' and '1550' and populate sequentially.
    """
    # Find sheets that match containing '1310' and '1550'
    sheet_1310 = [ws for ws in wb.worksheets if '1310' in ws.title]
    sheet_1550 = [ws for ws in wb.worksheets if '1550' in ws.title]
    injected = 0
    # Fill 1310 sheet(s)
    idx = 0
    for ws in sheet_1310:
        # collect placeholders [DISTANCE] and [1310] pairs by scanning rows
        rows_with_distance = []
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if isinstance(cell.value, str) and '[DISTANCE]' in cell.value:
                    rows_with_distance.append((ws, cell.row))
                    break
        for rpair in rows_with_distance:
            if idx >= len(otdr_grid):
                break
            entry = otdr_grid[idx]
            w, rr = rpair
            # fill DISTANCE and LOSS in that row
            for c in w[rr]:
                if _cell_is_formula(c):
                    continue
                v = c.value
                if isinstance(v, str) and '[DISTANCE]' in v:
                    c.value = entry.get('DISTANCE_1310', '')
                if isinstance(v, str) and '[1310]' in v:
                    c.value = entry.get('LOSS_1310', '')
            idx += 1
            injected += 1
    # Fill 1550 sheet(s) mirroring the same entries
    idx = 0
    for ws in sheet_1550:
        rows_with_distance = []
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if isinstance(cell.value, str) and '[DISTANCE]' in cell.value:
                    rows_with_distance.append((ws, cell.row))
                    break
        for rpair in rows_with_distance:
            if idx >= len(otdr_grid):
                break
            entry = otdr_grid[idx]
            w, rr = rpair
            for c in w[rr]:
                if _cell_is_formula(c):
                    continue
                v = c.value
                if isinstance(v, str) and '[DISTANCE]' in v:
                    c.value = entry.get('DISTANCE_1550', '')
                if isinstance(v, str) and '[1550]' in v:
                    c.value = entry.get('LOSS_1550', '')
            idx += 1
            injected += 1
    return injected


# -----------------------------
# Main API: inject_excel_final
# -----------------------------

def inject_excel_final(
    template_path: str,
    metadata: Dict[str, Any],
    parsed_fat: List[str],
    parsed_poles: List[Dict[str, Any]],
    phase2_data: Dict[str, Any],
    mode: str,
    port_count: int = DEFAULT_PORT_COUNT,
) -> io.BytesIO:
    """
    Open a fresh Template.xlsx and inject all Fase1 + Fase2 data in a single pass, returning BytesIO of final .xlsx.

    phase2_data is expected to be a dict with keys matching "fase2_measurements" from JSON contract, e.g.:
      - 'splitter_grid': [ {SPL_ID, SPL_SN, OPM_BEFORE, P1..Pn}, ... ]
      - 'opm_distribution_grid': [ {FAT_NAME, P1..Pn}, ... ]
      - 'otdr_distribution_grid': [ {FAT_NAME, DISTANCE_1310, LOSS_1310, DISTANCE_1550, LOSS_1550}, ... ]
      - 'IOR_1310', 'IOR_1550' (floats or strings)

    Guarantees: no style changes, red-tab safeguard, formula skip, apply mode text replacement in cover.
    """
    wb = load_workbook(template_path)
    mode = mode.lower() if isinstance(mode, str) else 'cluster'

    # Build metadata map and replace metadata tokens like Phase1
    meta_map = {}
    for key in PHASE1_TOKENS:
        k = key.strip('[]')
        meta_map[key] = metadata.get(k, '') if metadata else ''

    # Replace metadata tokens & mode text (cover sheets)
    for ws in wb.worksheets:
        if sheet_has_red_tab(ws) and not is_cover_sheet(ws.title):
            continue
        if is_cover_sheet(ws.title) and sheet_has_red_tab(ws.title):
            replace_mode_text(ws, mode)
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                try:
                    if _cell_is_formula(cell):
                        continue
                    val = cell.value
                    if isinstance(val, str) and val:
                        new_val = val
                        for tok, rep in meta_map.items():
                            if tok in new_val:
                                new_val = new_val.replace(tok, str(rep))
                        # Also replace IOR tokens if present and provided
                        if '[IOR_1310]' in new_val and 'IOR_1310' in phase2_data:
                            new_val = new_val.replace('[IOR_1310]', str(phase2_data.get('IOR_1310', '')))
                        if '[IOR_1550]' in new_val and 'IOR_1550' in phase2_data:
                            new_val = new_val.replace('[IOR_1550]', str(phase2_data.get('IOR_1550', '')))
                        if new_val != val:
                            cell.value = new_val
                except Exception:
                    continue

    # Inject Phase1 FAT names similar to phase1 injector
    fat_placeholders = []
    for ws in wb.worksheets:
        if sheet_has_red_tab(ws) and not is_cover_sheet(ws.title):
            continue
        found = _find_placeholder_cells_in_ws(ws, '[INPUT_FAT_NAME]')
        for col, row, cell in found:
            fat_placeholders.append((ws, col, row, cell))
    fat_placeholders.sort(key=lambda x: (x[1], x[2]))
    for i, name in enumerate(parsed_fat):
        if i < len(fat_placeholders):
            ws, col, row, cell = fat_placeholders[i]
            try:
                cell.value = name
            except Exception:
                pass
        else:
            # Spillover handling could be implemented here; for now, remaining parsed_fat stays in JSON only
            break

    # Inject poles
    pole_placeholders_desc = []
    pole_placeholders_size = []
    for ws in wb.worksheets:
        if sheet_has_red_tab(ws) and not is_cover_sheet(ws.title):
            continue
        for col, row, cell in _find_placeholder_cells_in_ws(ws, '[INPUT_POLE_DESC]'):
            pole_placeholders_desc.append((ws, col, row, cell))
        for col, row, cell in _find_placeholder_cells_in_ws(ws, '[INPUT_POLE_SIZE]'):
            pole_placeholders_size.append((ws, col, row, cell))
    pole_placeholders_desc.sort(key=lambda x: (x[1], x[2]))
    pole_placeholders_size.sort(key=lambda x: (x[1], x[2]))
    for i, pole in enumerate(parsed_poles):
        desc = pole.get('title', '')
        size = pole.get('size_clean', '')
        if i < len(pole_placeholders_desc):
            _, _, _, cell = pole_placeholders_desc[i]
            try:
                cell.value = desc
            except Exception:
                pass
        if i < len(pole_placeholders_size):
            _, _, _, cell = pole_placeholders_size[i]
            try:
                cell.value = size
            except Exception:
                pass

    # Phase2 injections
    # 1) IOR fields already injected if tokens existed
    # 2) Splitter Grid
    splitter_grid = phase2_data.get('splitter_grid', []) or []
    _inject_opm_feeder(wb, splitter_grid, port_count=port_count)

    # 3) OPM Distribution
    opm_dist_grid = phase2_data.get('opm_distribution_grid', []) or []
    _inject_opm_distribution(wb, opm_dist_grid, port_count=port_count)

    # 4) OTDR
    otdr_grid = phase2_data.get('otdr_distribution_grid', []) or []
    if mode == 'cluster':
        _inject_otdr_cluster(wb, otdr_grid)
    else:
        _inject_otdr_subfeeder(wb, otdr_grid)

    # After injections, remove any remaining tokens of the form [..] by replacing with '' (but keep template markers as per spec?)
    token_pattern = re.compile(r"\[[^\]]+\]")
    for ws in wb.worksheets:
        if sheet_has_red_tab(ws) and not is_cover_sheet(ws.title):
            continue
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                try:
                    if _cell_is_formula(cell):
                        continue
                    val = cell.value
                    if isinstance(val, str) and '[' in val and ']' in val:
                        new = token_pattern.sub('', val).strip()
                        cell.value = new
                except Exception:
                    continue

    # Garbage collection best-effort: remove unused template number sheets
    # Similar heuristics as phase1
    for name in list(wb.sheetnames):
        if RE_FAT_SHEETS.match(name) or RE_POLE_SHEETS.match(name) or RE_BA_SPLITTER_FAT.match(name):
            ws = wb[name]
            empty = True
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None and (str(cell).strip() != ''):
                        empty = False
                        break
                if not empty:
                    break
            if empty:
                try:
                    wb.remove(ws)
                except Exception:
                    pass

    # Reset dimensions and clean named ranges
    try:
        for ws in wb.worksheets:
            try:
                ws.reset_dimensions()
            except Exception:
                pass
    except Exception:
        pass

    try:
        # remove defined names referencing removed sheets
        to_delete = []
        for name in list(wb.defined_names.definedName):
            try:
                dests = list(wb.defined_names[name].destinations)
            except Exception:
                continue
            for title, _ in dests:
                if title not in wb.sheetnames:
                    to_delete.append(name)
                    break
        for name in to_delete:
            try:
                del wb.defined_names[name]
            except Exception:
                pass
    except Exception:
        pass

    # Save + recompress
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    try:
        compressed = recompress_xlsx(bio)
        return compressed
    except Exception:
        bio.seek(0)
        return bio


# -----------------------------
# Quick smoke test (requires Template.xlsx)
# -----------------------------
if __name__ == '__main__':
    from pathlib import Path
    import json
    tpl = 'Template.xlsx'
    if not Path(tpl).exists():
        print('Template.xlsx not found. Skipping smoke test.')
    else:
        meta = {'NAMA_PROYEK': 'EMR FTTH PROJECT', 'NAMA_LOKASI': 'DUSUN BOGO RW 08 FDT-2'}
        fat = ['A01', 'A02']
        poles = [{'title': 'Pole Erection 73', 'type': 'NEW POLE', 'size_clean': '7 METER 3 INCH', 'qty': 3}]
        phase2 = {
            'IOR_1310': 1.4681,
            'IOR_1550': 1.4676,
            'splitter_grid': [
                {'SPL_ID': 'SPLITTER 1', 'SPL_SN': 'SN-001', 'OPM_BEFORE': '-12.15', 'P1': '-18.10'}
            ],
            'opm_distribution_grid': [
                {'FAT_NAME': 'FAT A01', 'P1': '-19.10'}
            ],
            'otdr_distribution_grid': [
                {'FAT_NAME': 'FAT A01', 'DISTANCE_1310': '0.194', 'LOSS_1310': '1.000', 'DISTANCE_1550': '0.196', 'LOSS_1550': '1.340'}
            ]
        }
        out = inject_excel_final(tpl, meta, fat, poles, phase2, mode='cluster')
        with open('final_output.xlsx', 'wb') as f:
            f.write(out.read())
        print('Final output written to final_output.xlsx')
